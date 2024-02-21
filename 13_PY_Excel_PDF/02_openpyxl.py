import openpyxl
import os 

# os.chdir('C:\Users\\lrsor\\Desktop\\CURRENT-PROJECTS\\automate_python\\13_PY_Excel_PDF')

workbook = openpyxl.load_workbook('example.xlsx')

type(workbook)

# You can also call a specific sheet by passing in 
# the name: 

# sheet = workbook.get_sheet_by_name('sheetname')

# Alternatively, you can get all sheet names in the 
# workbook with:

print(workbook.sheetnames)
# ['Sheet1', 'Sheet2', 'Sheet3']

sheet = workbook['Sheet1']

cell = sheet['A1']
cell2 = sheet['B1']

print(cell.value, cell2.value)
# 2015-04-05 13:34:02 Apples

all = [str(sheet['A1'].value),sheet['B1'].value,sheet['C1'].value]

print(all)

# ['2015-04-05 13:34:02', 'Apples', 73]

sheet.cell(row=1, column=2)

for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)
# 1 Apples
# 2 Cherries
# 3 Pears
# 4 Oranges
# 5 Apples
# 6 Bananas
# 7 Strawberries
