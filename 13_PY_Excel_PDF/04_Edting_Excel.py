import os 
import openpyxl 


os.chdir('..\\13_PY_Excel_PDF\\')
# Create a brand new excel file with:
wb = openpyxl.Workbook()
print(wb.sheetnames)
# ['Sheet']

sheet = wb['Sheet']
print(sheet)
# <Worksheet "Sheet">
print(sheet['A1'].value == None)
# Above returns true since it's empty currently

sheet['A1'] = 42
sheet['A2'] = "Hello"
sheet['A3'] = "World"

print(sheet['A1'].value, sheet['A2'].value, sheet['A3'].value)
# 42 Hello World

# Right now above only exists in this program, to save 
# it to an actual file call:

wb.save('saveExample.xlsx')

# You can also add new sheets to the workbook:

sheet2 = wb.create_sheet()
print(wb.sheetnames)
# ['Sheet', 'Sheet1']
sheet2.title = 'Sheet2'
print(wb.sheetnames)
# ['Sheet', 'Sheet2']
wb.save('saveExample2.xlsx')

wb.create_sheet(index=0, title='index0Sheet')
# The above will insert a sheet at the beginngin of 
# the sheets collection, you can specify and index 
# to insert in the index=key above

wb.save('example3.xlsx')